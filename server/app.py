"""
Flask Backend Server
Handles file uploads, OCR processing, and PDF generation
"""

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import json
import shutil
import zipfile
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import base64
from io import BytesIO
from PIL import Image

from server.ocr_processor import OCRProcessor
from server.data_parser import DataParser
from server.pdf_filler import PDFFiller

app = Flask(__name__, static_folder='../web', static_url_path='')
CORS(app)

# Configuration
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'output')
TEMP_FOLDER = os.path.join(BASE_DIR, 'temp')
TEMPLATE_PDF = os.path.join(BASE_DIR, 'Target.pdf')

# Create folders if they don't exist
for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER, TEMP_FOLDER]:
    os.makedirs(folder, exist_ok=True)

# Allowed file extensions
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'tiff', 'bmp'}

# Initialize processors
ocr_processor = OCRProcessor()
data_parser = DataParser()
pdf_filler = PDFFiller(TEMPLATE_PDF)


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def cleanup_old_files():
    """Clean up old upload and temp files"""
    for folder in [UPLOAD_FOLDER, TEMP_FOLDER]:
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            try:
                if os.path.isfile(file_path):
                    # Delete files older than 1 hour
                    if os.path.getmtime(file_path) < (datetime.now().timestamp() - 3600):
                        os.unlink(file_path)
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")


def generate_thumbnail(file_path, max_size=(400, 400)):
    """Generate a base64-encoded thumbnail for preview"""
    try:
        # Handle PDFs - convert first page to image
        if file_path.lower().endswith('.pdf'):
            from pdf2image import convert_from_path
            images = convert_from_path(file_path, first_page=1, last_page=1)
            image = images[0] if images else None
        else:
            # Handle images
            image = Image.open(file_path)

        if image is None:
            return None

        # Auto-orient image based on EXIF data (fixes sideways phone photos)
        from PIL import ImageOps
        try:
            image = ImageOps.exif_transpose(image)
        except:
            pass

        # Create thumbnail
        image.thumbnail(max_size, Image.Resampling.LANCZOS)

        # Convert RGBA to RGB if necessary (for PNG with transparency)
        if image.mode in ('RGBA', 'LA', 'P'):
            # Create a white background
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode in ('RGBA', 'LA') else None)
            image = background

        # Convert to base64
        buffered = BytesIO()
        image.save(buffered, format="JPEG", quality=85)
        img_str = base64.b64encode(buffered.getvalue()).decode()

        return f"data:image/jpeg;base64,{img_str}"

    except Exception as e:
        print(f"Error generating thumbnail: {e}", flush=True)
        return None


@app.route('/')
def index():
    """Serve the main web interface"""
    return send_from_directory(app.static_folder, 'index.html')


@app.route('/api/upload', methods=['POST'])
def upload_files():
    """
    Handle file upload and OCR processing
    Returns extracted data for all uploaded files
    """
    try:
        # Clean up old files
        cleanup_old_files()

        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400

        files = request.files.getlist('files')

        if not files or files[0].filename == '':
            return jsonify({'error': 'No files selected'}), 400

        results = []

        for file in files:
            if file and allowed_file(file.filename):
                # Save uploaded file
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_filename = f"{timestamp}_{filename}"
                file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
                file.save(file_path)

                try:
                    # Process file with OCR
                    ocr_text = ocr_processor.process_file(file_path)

                    # DEBUG: Print raw OCR text
                    import sys
                    print("=" * 80, flush=True)
                    print(f"OCR TEXT FOR {filename}:", flush=True)
                    print("-" * 80, flush=True)
                    print(ocr_text, flush=True)
                    print("=" * 80, flush=True)

                    # Parse extracted text
                    extracted_data = data_parser.parse_text(ocr_text)

                    # DEBUG: Print extracted data
                    print("EXTRACTED DATA:", flush=True)
                    print(extracted_data, flush=True)
                    print("=" * 80, flush=True)

                    # Add source filename
                    extracted_data['source_filename'] = filename
                    extracted_data['ocr_text'] = ocr_text[:500]  # Include first 500 chars for debugging

                    # Generate thumbnail for preview
                    thumbnail = generate_thumbnail(file_path)

                    results.append({
                        'filename': filename,
                        'status': 'success',
                        'data': extracted_data,
                        'thumbnail': thumbnail
                    })

                except Exception as e:
                    results.append({
                        'filename': filename,
                        'status': 'error',
                        'error': str(e)
                    })

            else:
                results.append({
                    'filename': file.filename,
                    'status': 'error',
                    'error': 'Invalid file type'
                })

        return jsonify({
            'status': 'success',
            'results': results
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-single-pdf', methods=['POST'])
def generate_single_pdf():
    """
    Generate a single multi-page PDF with all declarations
    Returns a single PDF file
    """
    try:
        data_list = request.json.get('data', [])

        if not data_list:
            return jsonify({'error': 'No data provided'}), 400

        # Create unique output file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_path = os.path.join(OUTPUT_FOLDER, f'declarations_{timestamp}.pdf')

        # Generate single multi-page PDF
        pdf_filler.fill_single_multipage_pdf(data_list, pdf_path)

        return send_file(pdf_path, as_attachment=True, download_name=f'declarations_{timestamp}.pdf')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-pdfs', methods=['POST'])
def generate_pdfs():
    """
    Generate filled PDF declarations from extracted data
    Returns a ZIP file containing all PDFs
    """
    try:
        data_list = request.json.get('data', [])

        if not data_list:
            return jsonify({'error': 'No data provided'}), 400

        # Create unique output folder for this batch
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        batch_folder = os.path.join(TEMP_FOLDER, f'batch_{timestamp}')
        os.makedirs(batch_folder, exist_ok=True)

        # Generate PDFs (each data entry has its own seller_name)
        pdf_files = pdf_filler.fill_multiple_forms(data_list, batch_folder)

        # Create ZIP file
        zip_path = os.path.join(OUTPUT_FOLDER, f'SN_{timestamp}.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for pdf_file in pdf_files:
                zipf.write(pdf_file, os.path.basename(pdf_file))

        # Clean up individual PDF files
        shutil.rmtree(batch_folder)

        return send_file(zip_path, as_attachment=True, download_name=f'SN_{timestamp}.zip')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-excel', methods=['POST'])
def generate_excel():
    """
    Generate Excel file with all extracted data
    """
    try:
        data_list = request.json.get('data', [])

        if not data_list:
            return jsonify({'error': 'No data provided'}), 400

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inspection Data"

        # Define headers
        headers = ['Vendor', 'Source File', 'Stock #', 'Year', 'Make', 'Model', 'Body', 'Auto/Man',
                  'Colour', 'Engine No', 'VIN', 'Registration', 'Registration Expiry', 'Odometer']

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        for row_idx, data in enumerate(data_list, 2):
            ws.cell(row=row_idx, column=1, value=data.get('seller_name', ''))
            ws.cell(row=row_idx, column=2, value=data.get('source_filename', ''))
            ws.cell(row=row_idx, column=3, value=data.get('mta', ''))
            ws.cell(row=row_idx, column=4, value=data.get('year', ''))
            ws.cell(row=row_idx, column=5, value=data.get('make', ''))
            ws.cell(row=row_idx, column=6, value=data.get('model', ''))
            ws.cell(row=row_idx, column=7, value=data.get('type', ''))
            ws.cell(row=row_idx, column=8, value=data.get('transmission', ''))
            ws.cell(row=row_idx, column=9, value=data.get('color', ''))
            ws.cell(row=row_idx, column=10, value=data.get('engine_no', ''))
            ws.cell(row=row_idx, column=11, value=data.get('vin', ''))
            ws.cell(row=row_idx, column=12, value=data.get('reg', ''))
            ws.cell(row=row_idx, column=13, value=data.get('rego_expiry', ''))
            ws.cell(row=row_idx, column=14, value=data.get('odometer', ''))

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = os.path.join(OUTPUT_FOLDER, f'inspection_data_{timestamp}.xlsx')
        wb.save(excel_path)

        return send_file(excel_path, as_attachment=True, download_name=f'inspection_data_{timestamp}.xlsx')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    # Check if Tesseract is available
    tesseract_available = False
    tesseract_version = None
    try:
        import subprocess
        result = subprocess.run(['tesseract', '--version'],
                              capture_output=True, text=True, timeout=5)
        tesseract_available = result.returncode == 0
        if tesseract_available:
            tesseract_version = result.stdout.split('\n')[0]
    except Exception as e:
        tesseract_version = f"Error: {str(e)}"

    return jsonify({
        'status': 'healthy',
        'template_exists': os.path.exists(TEMPLATE_PDF),
        'tesseract_available': tesseract_available,
        'tesseract_version': tesseract_version
    })


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))

    print("=" * 80)
    print("Dec Filler Server Starting...")
    print(f"Upload folder: {UPLOAD_FOLDER}")
    print(f"Output folder: {OUTPUT_FOLDER}")
    print(f"Template PDF: {TEMPLATE_PDF}")
    print(f"Template exists: {os.path.exists(TEMPLATE_PDF)}")
    print(f"Port: {port}")
    print("=" * 80)
    print(f"\nServer running at http://localhost:{port}")
    print(f"Open your browser and navigate to http://localhost:{port}")
    print("\nPress Ctrl+C to stop the server")
    print("=" * 80)

    app.run(debug=False, host='0.0.0.0', port=port)
